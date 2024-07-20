from argparse import Namespace
import csv
import os
from collections import OrderedDict
import time
from typing import Any, Dict, Optional
import warnings
import yaml
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import uuid
import numpy as np

def get_experiment_name(model: str, backbone: str=None, seed=None):
    warnings.warn('get_experiment_name is deprecated. Use get_experiment_name_v2 instead.', DeprecationWarning)
    now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    model = f'{model}_{backbone}' if backbone else model
    if seed is not None:
        exp_name = os.path.join(model, str(seed), now)
    else:
        exp_name = os.path.join(model, 'Unseeded', now)
    return exp_name

# Oder of name abbreviations
# Higher elements should have higher granularity
MODEL_ABBREVIATIONS = {
    'resnet': 'ResNet',
    'swinv2_tiny': 'SwinV2Tiny',
    'swinv2_small': 'SwinV2Small',
    'swinv2_base': 'SwinV2Base',
    'swinv2_large': 'SwinV2Large',
    'swinv2': 'SwinV2',
    'swin': 'Swin',
}

def get_experiment_name_v2(args : Namespace) -> str:
    backbone = [v for k,v in MODEL_ABBREVIATIONS.items() if args.model.lower().startswith(k.lower())]
    # Use first matching abbreviation (highest granularity)
    backbone = backbone[0] if len(backbone) > 0 else args.model
    return _get_experiment_name_v2(method=args.method, 
                                   backbone=backbone, 
                                   classifier=args.classification_type, 
                                   head=args.head_type, 
                                   seed=args.seed,
                                   ds_split=args.ds_split,)

def _get_experiment_name_v2(name: Optional[str]=None,
                           method: Optional[str]=None, 
                           backbone: Optional[str]=None, 
                           classifier: Optional[str]=None,
                           head: Optional[str]=None,
                           seed:Optional[str|int]=None,
                           ds_split:Optional[str|int]=None,
                           include_time: bool=False,
                           *fields,
                           use_uuid: bool=True,
                           use_prefix: bool=False,
                           uuid_len: int=8) -> str:
    components = []
    # Append known components
    if name is not None and name != '':
        components.append(name)
    if method is not None and method != '':
        components.append(method)
    if backbone is not None and backbone != '':
        components.append(backbone)
    if classifier is not None and classifier != '':
        components.append(classifier)
    if head is not None and head != '':
        components.append(head)
    if seed is not None:
        components.append(str(seed))
    if ds_split is not None:
        components.append(str(ds_split))
    if include_time:
        now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
        components.append(now)
    # Append additional fields    
    components.extend(fields)
    # Append UUID
    if use_uuid:
        components.append(uuid.uuid4().hex[:uuid_len])
    if use_prefix:
        if name is not None and name != '':
            components = components[0]
        else:
            print('use_prefix is set to True but name is not provided. Ignoring use_prefix')
    return '_'.join(components)

def get_outdir(path, 
               *paths, 
               inc:bool=False, 
               use_uuid:bool=False, 
               uuid_len:int=8):
    outdir = os.path.join(path, *paths)
    if not os.path.exists(outdir):
        os.makedirs(outdir)
    elif inc:
        count = 1
        outdir_inc = outdir + '-' + str(count)
        while os.path.exists(outdir_inc):
            count = count + 1
            outdir_inc = outdir + '-' + str(count)
            assert count < 100
        outdir = outdir_inc
        os.makedirs(outdir)
    elif use_uuid:
        count = 1
        while os.path.exists(outdir):
            count = count + 1
            outdir = outdir + '-' + uuid.uuid4().hex[:uuid_len]
            assert count < 3
        os.makedirs(outdir)
    return outdir

def save_args(args, out_dir: str, filename: str = 'config.json'):
    # Cache the args as a text string to save them in the output dir later
    if filename[-5:] == '.yaml':
        args_text = yaml.safe_dump(args.__dict__, default_flow_style=False)
    elif filename[-5:] == '.json':
        args_text = json.dumps(args.__dict__, indent=4)
    else:
        filename = filename + ".yaml"
        
    with open(os.path.join(out_dir, filename), 'w') as f:
        f.write(args_text)

def update_summary(
        epoch: int | str,
        metrics: dict,
        root: str,
        write_header: bool = False,
        filename: str = 'results.csv'
) -> str:
    rowd = OrderedDict()
    if isinstance(epoch, str):
        rowd.update([('Phase', epoch)])
    else:
        rowd.update([('epoch', epoch)])
        
    rowd.update(metrics)
        
    if filename[-4:] != '.csv':
        filename = filename + '.csv'
    filename = os.path.join(root, filename)
    
    with open(filename, mode='a', newline='') as cf:
        dw = csv.DictWriter(cf, fieldnames=rowd.keys())
        if write_header: 
            dw.writeheader()
        dw.writerow(rowd)
        
    return filename
        
def convert_csv_to_excel(
    filepath: str
):
    excel_path = filepath[:-4] + '.xlsx' if filepath[-4:] == '.csv' else filepath + '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c).value = val
    wb.save(excel_path)
    
def create_or_append_excel(data: Dict[Any, dict], 
                           output_file: str, 
                           center: bool = False,
                           field_names: Optional[list | dict] = None) -> None:
    """Creates an excel or appends given data to an existing excel file. If the file doesn't exist, it creates a new one.
    The data is passed as a dictionary with the keys being the sheet names and the values being dictionaries with the data to be written.
    If the file doesn't exist, it creates a new one and creates appropriate headers.
    The header consists of the column number as the first row and `field_names` as the second row if given.
    The cells of the header are always centered. Otherwise centering can be toggled with the `center` parameter.

    Args:
        data (Dict[Any, dict]): 
            Data to be written to the excel file. The keys are the sheet names and the values are dictionaries with the data to be written.
        output_file (str): 
            Path to the output excel file.
        center (bool, optional): 
            Whether to center the content of the cells. Defaults to False.
        field_names (Optional[list  |  dict], optional): 
            Field names to be written when creating the sheets and headers. Only relevant if excel does not exist yet. Defaults to None.
    """
    if output_file[-5:] != '.xlsx':
        output_file += '.xlsx'
    # Check if the file exists
    if os.path.exists(output_file):
        # If the file exists, load it
        wb = openpyxl.load_workbook(output_file)
    else:
        # If the file doesn't exist, create a new workbook
        wb = openpyxl.Workbook()
        wb.save(output_file)
        if field_names is not None:
            if isinstance(field_names, list):
                field_names = {i: field_names[i] for i in range(len(field_names))}
            # Setup the sheets and headers
            create_or_append_excel(data={i: field_names for i in data.keys()},
                                    output_file=output_file, 
                                    center=True)
            wb = openpyxl.load_workbook(output_file)
        wb.remove(wb.active)  # Remove the default sheet

    for sheet_name, sheet_data in data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the next empty row
            next_row = ws.max_row + 1
        else:
            # If the sheet doesn't exist, create it and write headers
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1
            # Write field names in the first row
            for col, field_name in enumerate(sheet_data.keys(), start=1):
                cell = ws.cell(row=1, column=col, value=field_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            next_row = 2

        # Write the new data in the next empty row
        for col, value in enumerate(sheet_data.values(), start=1):
            # For NaN values, write -
            if isinstance(value, str) == False and np.isnan(value):
                value = '-'
            cell = ws.cell(row=next_row, column=col, value=value)
            if center:
                cell.alignment = Alignment(horizontal='center')

    # Save the workbook
    wb.save(output_file)
    
def export_reduced_conf(reduced_data: np.ndarray, 
                        filename: str ='reduced_confusion_matrix.xlsx',
                        sheet_name: Optional[str] = None):
    # Set default sheet name
    if sheet_name is None or sheet_name == '':
        sheet_name = 'Reduced Confusion Matrix'
    
    # Check if wb exists and set sheet
    if os.path.exists(filename):
        # Use existing workbook
        wb = openpyxl.load_workbook(filename)
        # Select or create the sheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(title=sheet_name)
    else:
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    # Write headers
    headers = ['Actual Class', 'Total Samples', 'Correct Predictions', 'Misclassifications']
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    max_misclassifications = 0

    # Write data
    for row, class_dict in enumerate(reduced_data, start=2):
        sheet.cell(row=row, column=1, value=class_dict['actual_class'])
        sheet.cell(row=row, column=2, value=class_dict['total_samples'])
        sheet.cell(row=row, column=3, value=class_dict['correct_predictions'])

        # Write misclassifications
        for col, (misclass_key, misclass_value) in enumerate(class_dict['misclassifications'].items(), start=4):
            cell = sheet.cell(row=row, column=col, value=f"{misclass_key} ({misclass_value})")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        max_misclassifications = max(max_misclassifications, len(class_dict['misclassifications']))

    # Merge misclassification header cells
    if max_misclassifications > 0:
        sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3 + max_misclassifications)

    # Center align all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    print(f"Excel file '{filename}' has been created successfully.")